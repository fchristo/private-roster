from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas


def main():
    r = Roster("../Jones_2019.xlsx")
    r.add_grades({"id": 3, "grades": [[2, 10], [5, 20], [12, 100]]})


class Roster(object):
    """A roster editor/ viewer for poor Mrs.Jones"""

    def get_student_names(self):
        """Gets and returns a complete list of all student's names"""
        student_names = []
        student_workbook = load_workbook(self.filename)
        sheet = student_workbook.active
        class_dataframe = pandas.DataFrame(sheet.values)

        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            for cell in row:
                student_names.append(class_dataframe.loc[cell, 1] + ' ' + class_dataframe.loc[cell, 2])

        return student_names

    def get_student(self, student_identifier):
        """Reads excel for student, creates the student if they don't exist, returns the student's ID and grades"""
        student_workbook = load_workbook(self.filename)
        sheet = student_workbook.active
        class_dataframe = pandas.DataFrame(sheet.values)

        if isinstance(student_identifier, int):
            student_id = student_identifier

            try:
                student_sheet = student_workbook["Student_" + student_identifier.__str__()]
            except KeyError as e:
                print(e)
                raise Exception('ID not found. If you would like to add a student, enter a full name. ' +
                                'Otherwise, please enter the name or ID of an existing student.')

            grade_list = []

            for row in student_sheet.iter_rows(min_row=6, min_col=2, values_only=True):
                for cell in row:
                    grade_list.append(cell)

            grades = pandas.Series(grade_list)

            student = {'id': student_id, 'grades': grades}

            print(student['grades'][4])

            return student
        elif isinstance(student_identifier, str):
            student_name = student_identifier.split(' ')

            if class_dataframe[2].isin(student_name).any():  # check if student name exists
                student_dataframe = pandas.DataFrame(
                    class_dataframe.loc[class_dataframe[2].isin(student_name)])  # get row of student based on name
                student_id = student_dataframe[0].values[0]  # get student id from row
                student_sheet = student_workbook["Student_" + student_id.__str__()]
            else:
                student_id = 1
                # update all other students ID's in Roster and subsheets
                for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    for cell in row:
                        next_cell = cell + 1
                        next_cell_str = next_cell.__str__()
                        current_sheet = student_workbook["Student_" + str(cell)]
                        current_sheet["B1"] = next_cell
                        sheet["A" + next_cell_str] = next_cell
                        sheet["D" + next_cell_str] = "=Student_" + next_cell_str + "!B3"

                # rename the student sheets (backwards as to avoid two files having the same name)
                for student_num in range(len(self.get_student_names()), 0, -1):
                    student_workbook["Student_" +
                                     student_num.__str__()].title = "Student_" + (student_num + 1).__str__()

                # move all other students down one row
                sheet.move_range("A2:D8", rows=1)

                # make new student Student_1
                sheet["A2"] = student_id
                sheet["B2"] = student_name[0]
                sheet["C2"] = student_name[1]
                sheet["D2"] = "=Student_1!B3"
                student_workbook.create_sheet('Student_1')
                student_sheet = student_workbook['Student_1']
                self._write_default_fields(student_sheet, student_identifier)
                self.save("Jones_2019_Updated.xlsx", student_workbook)

            grade_list = []

            # compile a list of the student's grades
            for row in student_sheet.iter_rows(min_row=6, min_col=2, values_only=True):
                for cell in row:
                    grade_list.append(cell)

            grades = pandas.Series(grade_list)

            student = {'id': student_id, 'grades': grades}

            return student
        else:
            raise Exception('Please enter either a full name or a student ID')

    def delete_student(self, student_identifier):
        pandas.read_excel(self.filename)

    def save(self, output_filename: str, workbook=None):
        """Save the passed in workbook, or save loaded workbook if not passed."""
        if workbook is None:
            workbook = load_workbook(self.filename)

        workbook.save(output_filename)

    def class_average(self, workbook=None):
        """Read each student's sheet, get the GPA of each student, then return the GPA of the class"""
        if workbook is None:
            workbook = load_workbook(self.filename)

        sheet = workbook.active
        grades_total = 0
        students_num = self.get_student_names().__len__()

        # loop through all the student sheets
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            for cell in row:
                current_sheet = workbook["Student_" + str(cell)]
                grades_list = []
                # loop through all the grades in each student's sheet
                for grade_row in current_sheet.iter_rows(min_row=6, min_col=2, values_only=True):
                    for grade_cell in grade_row:
                        grades_list.append(grade_cell)

                grades_total = grades_total + pandas.Series(grades_list).mean()

        avg_grade = grades_total / students_num

        return avg_grade

    def add_grades(self, student: dict):
        """Add grades to a student's sheet"""
        student_workbook = load_workbook(self.filename)

        if student["id"]:
            sheet = student_workbook["Student_" + student["id"].__str__()]
            student_dataframe = pandas.DataFrame(sheet.values)
            grades_dataframe = pandas.DataFrame(student["grades"])

            # iterate over all assignments in the student's sheet
            for row in sheet.iter_rows(min_row=6, max_col=1, values_only=True):
                for cell in row:
                    # iterate over each assignment passed in
                    for assignment in list(zip(*student["grades"]))[0]:
                        # if assignment exists, update grade value
                        if cell == assignment:
                            student_dataframe.iloc[cell + 4][1] = \
                                grades_dataframe[grades_dataframe[0] == cell].iat[0, 1]
                        elif assignment > student_dataframe[0].iloc[-1]:  # if the assignment does not exist, add it
                            assignment_dataframe = grades_dataframe.loc[grades_dataframe[0] == assignment]
                            student_dataframe = student_dataframe.append(assignment_dataframe, ignore_index=True)

            # Convert from pandas and apply changes to the actual worksheet
            rows = dataframe_to_rows(student_dataframe, index=False, header=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            self.save("Jones_2019_Updated.xlsx", student_workbook)
        else:
            raise Exception("Please enter a valid student ID")

    @staticmethod
    def _write_default_fields(student_sheet, student_identifier):
        """Fill out the default fields for a new student's sheet"""
        rows = (
            ("Student ID", 1),
            ("Name", student_identifier),
            ("Grade", "=SUM(B6:B15)/10"),
            ("", ""),
            ("Assignment", "Grade"),
        )

        for row in rows:
            student_sheet.append(row)

    def __init__(self, filename):
        self.filename = filename
        pass


if __name__ == "__main__":
    main()
