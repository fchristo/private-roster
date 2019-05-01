import unittest
import pandas
from openpyxl import load_workbook
from roster import Roster


class TestRoster(unittest.TestCase):

    def test_read_roster(self):
        with Roster("Jones_2019.xlsx") as r:
            student_names = r.get_student_names()
            self.assertTrue(len(student_names) == 7)
            self.assertTrue("Robert Waters" in student_names)

            catherine = r.get_student("Catherine Hitchens")
            self.assertTrue(catherine["id"] == 3)
            self.assertTrue(isinstance(catherine["grades"], pandas.Series))
            self.assertTrue(len(catherine["grades"]) == 10)
            self.assertTrue(catherine["grades"][4] == 86)

            self.assertTrue(r.class_average() == 614.1/7)

    def test_write_roster(self):
        with Roster("Jones_2019.xlsx") as r:
            john = r.get_student("Johnny Carson")
            for assignment, grade in [(3, 90), (6, 94), (9, 92)]:
                john["grades"][assignment] = grade
            self.assertTrue(r.class_average() == 618.7/7)
            r.save("Jones_2019_Updated.xlsx")

        wb = load_workbook("Jones_2019_Updated.xlsx")
        self.assertTrue(wb.get_sheet_by_name("Student_1")["B10"].value == 94)
        wb.close()

    def test_delete_roster_student(self):
        student_count = 0
        with Roster("Jones_2019.xlsx") as r:
            student_count = len(r.get_student_names())
            self.assertTrue(student_count == 7)
            self.assertTrue(r.get_student("William Thomas")["id"] == 5)
            r.delete_student("Allen Dalton")
            student_count = len(r.get_student_names())
            self.assertTrue(student_count == 6)
            self.assertTrue(r.get_student("William Thomas")["id"] == 4)
            r.save("Jones_2019_Reduced.xlsx")

        wb = load_workbook("Jones_2019_Reduced.xlsx")
        sheet_names = wb.get_sheet_names()
        self.assertTrue(len(sheet_names) == 7)
        self.assertTrue(sheet_names[0] == "Roster")
        self.assertTrue(sheet_names[-1] == "Student_6")
        self.assertTrue(wb.get_sheet_by_name("Student_3")["B7"] == 92)
        wb.close()

if __name__ == "__main__":
    unittest.main()
