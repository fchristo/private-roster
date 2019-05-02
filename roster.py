"""
roster.py
====================================
A roster editor/ viewer for poor Mrs.Jones
"""

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas


def main():
    print("To execute this file's functions, use invoke <taskname>")


class Roster(object):
    """
    A roster object
    """

    def __enter__(self):
        self.roster = Roster(self.filename)
        return self

    def get_student_names(self):
        """
            Return a a list of all student names
        """
        class_dataframe = pandas.DataFrame(self.sheet.values)
        student_names = []

        for row in self.sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            for cell in row:
                student_names.append(
                    class_dataframe.loc[cell, 1] + " " + class_dataframe.loc[cell, 2]
                )
        return student_names

    def get_student(self, student_identifier):
        """
            Reads excel for student, creates the student if they don't exist, returns the student's ID and grades

            Parameters
            ---------
            student_identifier
                Either a full name as a string, or a student ID as an int
        """
        class_dataframe = pandas.DataFrame(self.sheet.values)
        if isinstance(student_identifier, int):
            student_id = student_identifier

            try:
                student_sheet = self.student_workbook[
                    "Student_" + student_identifier.__str__()
                ]
            except KeyError as e:
                print(e)
                raise Exception(
                    "ID not found. If you would like to add a student, enter a full name. "
                    + "Otherwise, please enter the name or ID of an existing student."
                )

            grade_list = []

            for row in student_sheet.iter_rows(min_row=6, min_col=2, values_only=True):
                for cell in row:
                    grade_list.append(cell)

            grades = pandas.Series(grade_list)

            student = {"id": student_id, "grades": grades}

            return student
        elif isinstance(student_identifier, str):
            student_name = student_identifier.split(" ")
            if (
                class_dataframe[2].isin(student_name).any()
            ):  # check if student name exists
                student_dataframe = pandas.DataFrame(
                    # get row of student based on name
                    class_dataframe.loc[class_dataframe[2].isin(student_name)]
                )
                student_id = student_dataframe[0].values[0]  # get student id from row
                student_sheet = self.student_workbook["Student_" + student_id.__str__()]
            else:  # create a student
                student_id = 1
                # update all other students ID's in Roster and subsheets
                for row in self.sheet.iter_rows(min_row=2, max_col=1, values_only=True):
                    for cell in row:
                        next_cell = cell + 1
                        next_cell_str = next_cell.__str__()
                        current_sheet = self.student_workbook["Student_" + str(cell)]
                        current_sheet["B1"] = next_cell
                        self.sheet["A" + next_cell_str] = next_cell
                        self.sheet["D" + next_cell_str] = (
                            "=Student_" + next_cell_str + "!B3"
                        )

                # rename the student sheets (backwards as to avoid two files having the same name)
                for student_num in range(len(self.get_student_names()), 0, -1):
                    self.student_workbook["Student_" + student_num.__str__()].title = (
                        "Student_" + (student_num + 1).__str__()
                    )

                # move all other students down one row
                self.sheet.move_range("A2:D8", rows=1)

                # make new student Student_1
                self.sheet["A2"] = student_id
                self.sheet["B2"] = student_name[0]
                self.sheet["C2"] = student_name[1]
                self.sheet["D2"] = "=Student_1!B3"
                self.student_workbook.create_sheet("Student_1")
                student_sheet = self.student_workbook["Student_1"]
                self._write_default_fields(student_sheet, student_identifier)
                updated_file = self.filename.split(".xlsx")
                self.save(updated_file[0] + "_Updated.xlsx")

            grade_list = []

            # compile a list of the student's grades
            for row in student_sheet.iter_rows(min_row=6, min_col=2, values_only=True):
                for cell in row:
                    grade_list.append(cell)

            grades = pandas.Series(grade_list)

            student = {"id": student_id, "grades": grades}

            return student
        else:
            raise Exception("Please enter either a full name or a student ID")

    def delete_student(self, student_identifier):
        """
            Deletes a student from the workbook

            Parameters
            ---------
            student_identifier
                Either a full name as a string, or a student ID as an int
        """
        class_dataframe = pandas.DataFrame(self.sheet.values)
        if isinstance(student_identifier, int):
            self._do_delete(self.student_workbook, student_identifier)
        elif isinstance(student_identifier, str):
            student_name = student_identifier.split(" ")
            if (
                class_dataframe[2].isin(student_name).any()
            ):  # check if student name exists
                student_dataframe = pandas.DataFrame(
                    # get row of student based on name
                    class_dataframe.loc[class_dataframe[2].isin(student_name)]
                )
                student_id = student_dataframe[0].values[0]  # get student id from row
                self._do_delete(self.student_workbook, student_id)
            else:
                raise Exception(
                    "Please enter an existing student, or use get_student('student') to create one"
                )
        else:
            raise Exception(
                "To delete a student, provide a valid student ID or full name"
            )

        updated_file = self.filename.split(".xlsx")
        self.save(updated_file[0] + "_Updated.xlsx")

    def save(self, output_filename: str, workbook=None):
        """
            Save the passed in workbook, or save loaded workbook if not passed.

            Parameters
            ---------
            output_filename
                Name of the file to save to

            workbook
                Workbook to save
        """
        if workbook is None:
            workbook = self.student_workbook

        workbook.save(output_filename)

    def class_average(self, workbook=None):
        """
            Read each student's sheet, get the GPA of each student, then return the GPA of the class

            Parameters
            ---------
            workbook
                Workbook to get students from
        """
        if workbook is None:
            workbook = self.student_workbook

        sheet = workbook.active
        grades_total = 0
        students_num = self.get_student_names().__len__()

        # loop through all the student sheets
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            for cell in row:
                current_sheet = workbook["Student_" + str(cell)]
                grades_list = []
                # loop through all the grades in each student's sheet
                for grade_row in current_sheet.iter_rows(
                    min_row=6, min_col=2, values_only=True
                ):
                    for grade_cell in grade_row:
                        grades_list.append(grade_cell)

                grades_total = grades_total + pandas.Series(grades_list).mean()

        avg_grade = grades_total / students_num

        return avg_grade

    def add_grades(self, student: dict):
        """
            Add grades to a student's sheet

            Parameters
            ---------
            student
                dictionary containing student ID and a list of assignments/grades.
                Input example: {"id": 1, "grades": [(3, 90), (12, 100), (1, 10)]}
        """
        if student["id"]:
            sheet = self.student_workbook["Student_" + student["id"].__str__()]
            student_dataframe = pandas.DataFrame(sheet.values)
            grades_dataframe = pandas.DataFrame(student["grades"])

            # iterate over all assignments in the student's sheet
            for row in sheet.iter_rows(min_row=6, max_col=1, values_only=True):
                for cell in row:
                    # iterate over each assignment passed in
                    for assignment in list(zip(*student["grades"]))[0]:
                        # if assignment exists, update grade value
                        if cell == assignment:
                            student_dataframe.iloc[cell + 4][1] = grades_dataframe[
                                grades_dataframe[0] == cell
                            ].iat[0, 1]
                        elif (
                            assignment > student_dataframe[0].iloc[-1]
                        ):  # if the assignment does not exist, add it
                            assignment_dataframe = grades_dataframe.loc[
                                grades_dataframe[0] == assignment
                            ]
                            student_dataframe = student_dataframe.append(
                                assignment_dataframe, ignore_index=True
                            )

            # Convert from pandas and apply changes to the actual worksheet
            rows = dataframe_to_rows(student_dataframe, index=False, header=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

            updated_file = self.filename.split(".xlsx")
            self.save(updated_file[0] + "_Updated.xlsx")
        else:
            raise Exception("Please enter a valid student ID")

    @staticmethod
    def _do_delete(workbook, student_id: int):
        """
            Logic to delete student from worksheet. Extrapolated to keep DRY

            Parameters
            ---------
            workbook
                the workbook to delete a student from

            student_id
                The student ID number
        """
        workbook.remove(workbook["Student_" + str(student_id)])
        # Rename all the student sheets to be one less than they were
        for sheet in workbook.sheetnames:
            sheet_num = sheet.title()
            sheet_num = sheet_num.split("_")

            # check that student sheet is after deleted sheet
            if sheet_num[0] != "Roster" and int(sheet_num[1]) > student_id:
                workbook[sheet].cell(row=1, column=2).value = int(sheet_num[1]) - 1
                workbook[sheet].title = "Student_" + (int(sheet_num[1]) - 1).__str__()

        workbook["Roster"].delete_rows(student_id + 1, 1)

        # loop through each student's ID and decrease it by one if after deleted student
        for row in workbook["Roster"].iter_rows(min_row=2, max_col=1):
            for cell in row:
                if cell.value > student_id:
                    workbook["Roster"].cell(cell.row, cell.column).value = (
                        cell.value - 1
                    )

        return workbook

    @staticmethod
    def _write_default_fields(student_sheet, student_identifier: str):
        """
            Logic to fill out the default fields a newly created student sheet should have.

            Parameters
            ---------
            student_sheet
                the student sheet we are writing to

            student_identifier
                The student name
        """
        rows = (
            ("Student ID", 1),
            ("Name", student_identifier),
            ("Grade", "=SUM(B6:B15)/10"),
            ("", ""),
            ("Assignment", "Grade"),
        )

        for row in rows:
            student_sheet.append(row)

    def __init__(self, filename):
        self.filename = filename
        self.student_workbook = load_workbook(filename)
        self.sheet = self.student_workbook.active
        self.roster = None
        pass

    def __exit__(self, exception_type, exception_value, traceback):
        print(exception_type, exception_value, traceback)


if __name__ == "__main__":
    main()
